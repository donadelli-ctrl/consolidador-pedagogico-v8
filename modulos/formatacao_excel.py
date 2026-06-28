from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment
)

from openpyxl.utils import get_column_letter


# ==========================================================
# FORMATAÇÃO DAS ABAS
# ==========================================================

def aplicar_cores(ws):

    # ======================================================
    # CABEÇALHO
    # ======================================================

    azul = PatternFill(

        "solid",

        fgColor="1F4E79"

    )

    for celula in ws[1]:

        celula.font = Font(

            color="FFFFFF",

            bold=True

        )

        celula.fill = azul

        celula.alignment = Alignment(

            horizontal="center",

            vertical="center"

        )

    # ======================================================
    # FORMATAÇÃO DAS CÉLULAS
    # ======================================================

    for linha in ws.iter_rows(min_row=2):

        for celula in linha:

            valor = celula.value

            celula.alignment = Alignment(

                horizontal="center",

                vertical="center"

            )

            # ----------------------------------------------
            # PERCENTUAIS
            # ----------------------------------------------

            if isinstance(valor, (int, float)):

                if 0 <= valor <= 1:

                    celula.number_format = "0.0%"

            texto = str(valor).strip()

            # ----------------------------------------------
            # EVOLUÇÃO
            # ----------------------------------------------

            if texto == "Melhorou":

                celula.font = Font(

                    color="008000",

                    bold=True

                )

            elif texto == "Manteve":

                celula.font = Font(

                    color="C9A000",

                    bold=True

                )

            elif texto == "Piorou":

                celula.font = Font(

                    color="FF0000",

                    bold=True

                )

            elif texto == "Sem Comparação":

                celula.font = Font(

                    color="808080",

                    italic=True

                )

            # ----------------------------------------------
            # SITUAÇÃO
            # ----------------------------------------------

            elif texto == "Adequado":

                celula.font = Font(

                    color="008000",

                    bold=True

                )

            elif texto == "Acompanhamento":

                celula.font = Font(

                    color="C9A000",

                    bold=True

                )

            elif texto == "Atenção":

                celula.font = Font(

                    color="FF0000",

                    bold=True

                )

            # ----------------------------------------------
            # NÍVEIS
            # ----------------------------------------------

            elif texto == "Abaixo do Básico":

                celula.font = Font(

                    color="FF0000",

                    bold=True

                )

            elif texto == "Básico":

                celula.font = Font(

                    color="C9A000",

                    bold=True

                )

            elif texto == "Proficiente":

                celula.font = Font(

                    color="0000FF",

                    bold=True

                )

    # ======================================================
    # CONGELAR PRIMEIRA LINHA
    # ======================================================

    ws.freeze_panes = "A2"

    # ======================================================
    # FILTRO
    # ======================================================

    ws.auto_filter.ref = ws.dimensions

    # ======================================================
    # AJUSTAR LARGURA DAS COLUNAS
    # ======================================================

    for coluna in ws.columns:

        tamanho = 0

        letra = get_column_letter(

            coluna[0].column

        )

        for celula in coluna:

            try:

                tamanho = max(

                    tamanho,

                    len(

                        str(

                            celula.value

                        )

                    )

                )

            except:

                pass

        ws.column_dimensions[

            letra

        ].width = min(

            tamanho + 3,

            40

        )
